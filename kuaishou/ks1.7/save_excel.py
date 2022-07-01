import openpyxl



class SaveExecl:

    '''
    保存excel文件
    '''

    def __init__(self, file_name, sheet_name='sheet1'):
        self.wb = openpyxl.Workbook()
        self.ws = self.wb.create_sheet(index=0, title=sheet_name)
        self.file_name = file_name + '.xlsx' if file_name.split('.')[-1] != 'xlsx' else file_name


    def save_to_excel(self,row_num,column_num,data):
        """
        将信息保存到excel表中;
        [data]为要保存的数据;
        """
        self.ws.cell(row=row_num, column=column_num, value=data)

    def sucess_sing(self,sign):
        if sign is True:
            self.wb.save(filename=self.file_name)